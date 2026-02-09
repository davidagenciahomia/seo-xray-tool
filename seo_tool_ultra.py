import streamlit as st
import requests
from bs4 import BeautifulSoup
import pandas as pd
import re
from collections import Counter
import trafilatura
import urllib3
from datetime import datetime
import json
from textstat import flesch_reading_ease
import numpy as np
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.cluster import KMeans
from urllib.parse import urlparse
import time
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

# ==========================================
# CONFIGURACIÓN INICIAL
# ==========================================
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
st.set_page_config(
    page_title="SERP X-Ray 360 | Professional SEO Analysis", 
    layout="wide", 
    page_icon="📊",
    initial_sidebar_state="expanded"
)

# ==========================================
# CSS CORPORATIVO PROFESIONAL
# ==========================================
st.markdown("""
<style>
    @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap');
    
    :root {
        --primary: #FF6B00;
        --primary-dark: #E55D00;
        --secondary: #2563eb;
        --success: #059669;
        --warning: #d97706;
        --danger: #dc2626;
        --bg-dark: #0a0e27;
        --bg-secondary: #151935;
        --bg-tertiary: #1e2139;
        --text-primary: #f8fafc;
        --text-secondary: #94a3b8;
        --border: #2d3250;
    }
    
    * {
        font-family: 'Inter', -apple-system, BlinkMacSystemFont, sans-serif;
    }
    
    .stApp {
        background: var(--bg-dark);
    }
    
    [data-testid="stSidebar"] {
        background: var(--bg-secondary);
        border-right: 1px solid var(--border);
    }
    
    [data-testid="stSidebar"] .block-container {
        padding-top: 2rem;
    }
    
    /* HEADER PRINCIPAL */
    .main-header {
        background: linear-gradient(135deg, var(--bg-secondary) 0%, var(--bg-tertiary) 100%);
        padding: 2rem;
        border-radius: 8px;
        border: 1px solid var(--border);
        margin-bottom: 2rem;
    }
    
    .main-header h1 {
        font-size: 2rem !important;
        font-weight: 700 !important;
        color: var(--text-primary) !important;
        margin: 0 !important;
        letter-spacing: -0.02em;
    }
    
    .main-header .subtitle {
        color: var(--text-secondary);
        font-size: 0.95rem;
        margin-top: 0.5rem;
        font-weight: 400;
    }
    
    /* TÍTULOS */
    h1, h2, h3, h4, h5, h6 {
        color: var(--text-primary) !important;
        font-weight: 600 !important;
        letter-spacing: -0.01em;
    }
    
    h2 {
        font-size: 1.5rem !important;
        margin-top: 2rem !important;
    }
    
    h3 {
        font-size: 1.25rem !important;
        margin-top: 1.5rem !important;
    }
    
    /* MÉTRICAS */
    [data-testid="stMetric"] {
        background: var(--bg-secondary);
        padding: 1.25rem;
        border-radius: 8px;
        border: 1px solid var(--border);
        transition: all 0.2s ease;
    }
    
    [data-testid="stMetric"]:hover {
        border-color: var(--primary);
        transform: translateY(-2px);
    }
    
    [data-testid="stMetric"] label {
        color: var(--text-secondary) !important;
        font-size: 0.8rem !important;
        font-weight: 500 !important;
        text-transform: uppercase;
        letter-spacing: 0.05em;
    }
    
    [data-testid="stMetric"] [data-testid="stMetricValue"] {
        color: var(--text-primary) !important;
        font-size: 1.75rem !important;
        font-weight: 700 !important;
    }
    
    /* TABS */
    .stTabs [data-baseweb="tab-list"] {
        gap: 4px;
        background: var(--bg-secondary);
        padding: 4px;
        border-radius: 8px;
        border: 1px solid var(--border);
    }
    
    .stTabs [data-baseweb="tab"] {
        height: 44px;
        padding: 0 20px;
        background: transparent;
        border-radius: 6px;
        color: var(--text-secondary);
        font-weight: 500;
        font-size: 0.9rem;
        border: none;
    }
    
    .stTabs [aria-selected="true"] {
        background: var(--primary) !important;
        color: white !important;
    }
    
    /* BOTONES */
    .stButton button {
        background: var(--primary);
        color: white;
        border: none;
        padding: 0.625rem 1.5rem;
        border-radius: 6px;
        font-weight: 600;
        font-size: 0.9rem;
        transition: all 0.2s ease;
        letter-spacing: 0.01em;
    }
    
    .stButton button:hover {
        background: var(--primary-dark);
        transform: translateY(-1px);
        box-shadow: 0 4px 12px rgba(255, 107, 0, 0.3);
    }
    
    /* INPUTS */
    .stTextInput input, .stNumberInput input, .stSelectbox select {
        background: var(--bg-tertiary) !important;
        border: 1px solid var(--border) !important;
        border-radius: 6px !important;
        color: var(--text-primary) !important;
        font-size: 0.9rem !important;
        padding: 0.625rem !important;
    }
    
    .stTextInput input:focus, .stNumberInput input:focus, .stSelectbox select:focus {
        border-color: var(--primary) !important;
        box-shadow: 0 0 0 1px var(--primary) !important;
    }
    
    /* EXPANDERS */
    .streamlit-expanderHeader {
        background: var(--bg-secondary);
        border: 1px solid var(--border);
        border-radius: 8px;
        padding: 1rem 1.25rem;
        font-weight: 500;
        color: var(--text-primary);
        font-size: 0.95rem;
    }
    
    .streamlit-expanderHeader:hover {
        border-color: var(--primary);
    }
    
    .streamlit-expanderContent {
        background: var(--bg-tertiary);
        border: 1px solid var(--border);
        border-top: none;
        border-radius: 0 0 8px 8px;
        padding: 1.25rem;
    }
    
    /* TABLAS */
    .dataframe {
        font-size: 0.85rem !important;
    }
    
    .dataframe thead tr th {
        background: var(--bg-secondary) !important;
        color: var(--text-secondary) !important;
        font-weight: 600 !important;
        text-transform: uppercase;
        font-size: 0.75rem !important;
        letter-spacing: 0.05em;
        padding: 0.875rem !important;
        border-bottom: 2px solid var(--primary) !important;
    }
    
    .dataframe tbody tr {
        background: var(--bg-tertiary) !important;
        border-bottom: 1px solid var(--border) !important;
    }
    
    .dataframe tbody tr:hover {
        background: var(--bg-secondary) !important;
    }
    
    .dataframe tbody td {
        color: var(--text-primary) !important;
        padding: 0.875rem !important;
    }
    
    /* ALERTS */
    .stAlert {
        background: var(--bg-secondary);
        border: 1px solid var(--border);
        border-radius: 8px;
        padding: 1rem;
        color: var(--text-primary);
    }
    
    /* INFO BOXES */
    .info-box {
        background: linear-gradient(135deg, rgba(37, 99, 235, 0.1) 0%, rgba(37, 99, 235, 0.05) 100%);
        border-left: 3px solid var(--secondary);
        padding: 1rem 1.25rem;
        border-radius: 6px;
        margin: 1rem 0;
        font-size: 0.9rem;
        line-height: 1.6;
        color: var(--text-primary);
    }
    
    .success-box {
        background: linear-gradient(135deg, rgba(5, 150, 105, 0.1) 0%, rgba(5, 150, 105, 0.05) 100%);
        border-left: 3px solid var(--success);
        padding: 1rem 1.25rem;
        border-radius: 6px;
        margin: 1rem 0;
        font-size: 0.9rem;
        line-height: 1.6;
        color: var(--text-primary);
    }
    
    .warning-box {
        background: linear-gradient(135deg, rgba(217, 119, 6, 0.1) 0%, rgba(217, 119, 6, 0.05) 100%);
        border-left: 3px solid var(--warning);
        padding: 1rem 1.25rem;
        border-radius: 6px;
        margin: 1rem 0;
        font-size: 0.9rem;
        line-height: 1.6;
        color: var(--text-primary);
    }
    
    /* BADGES */
    .badge {
        display: inline-block;
        padding: 0.25rem 0.625rem;
        border-radius: 4px;
        font-size: 0.75rem;
        font-weight: 600;
        letter-spacing: 0.02em;
        margin: 0.125rem;
    }
    
    .badge-success {
        background: rgba(5, 150, 105, 0.2);
        color: #10b981;
        border: 1px solid rgba(5, 150, 105, 0.3);
    }
    
    .badge-warning {
        background: rgba(217, 119, 6, 0.2);
        color: #f59e0b;
        border: 1px solid rgba(217, 119, 6, 0.3);
    }
    
    .badge-danger {
        background: rgba(220, 38, 38, 0.2);
        color: #ef4444;
        border: 1px solid rgba(220, 38, 38, 0.3);
    }
    
    .badge-info {
        background: rgba(37, 99, 235, 0.2);
        color: #3b82f6;
        border: 1px solid rgba(37, 99, 235, 0.3);
    }
    
    /* DIVISORES */
    hr {
        border: none;
        height: 1px;
        background: var(--border);
        margin: 2rem 0;
    }
    
    /* TOOLTIPS DISCRETOS */
    .tooltip-icon {
        display: inline-block;
        width: 16px;
        height: 16px;
        background: var(--text-secondary);
        color: var(--bg-dark);
        border-radius: 50%;
        text-align: center;
        line-height: 16px;
        font-size: 10px;
        font-weight: 700;
        cursor: help;
        margin-left: 4px;
        opacity: 0.6;
        transition: opacity 0.2s;
    }
    
    .tooltip-icon:hover {
        opacity: 1;
    }
    
    /* SIDEBAR */
    [data-testid="stSidebar"] h1,
    [data-testid="stSidebar"] h2,
    [data-testid="stSidebar"] h3 {
        font-size: 1.1rem !important;
        margin-bottom: 1rem !important;
    }
    
    [data-testid="stSidebar"] label {
        color: var(--text-secondary) !important;
        font-size: 0.8rem !important;
        font-weight: 500 !important;
        text-transform: uppercase;
        letter-spacing: 0.05em;
    }
    
    /* PROGRESS BAR */
    .stProgress > div > div > div > div {
        background: var(--primary);
    }
</style>
""", unsafe_allow_html=True)

# ==========================================
# TOOLTIPS SIMPLIFICADOS
# ==========================================
TOOLTIPS = {
    "word_count": "Average word count across top 10. Aim for 20% above this number.",
    "internal_links": "Internal links help distribute page authority and improve site structure.",
    "images": "Visual content improves engagement and time on page.",
    "da_proxy": "Domain Authority proxy based on backlink profile indicators.",
    "h2_critical": "H2 headings that appear in more than 50% of competing pages.",
}

def show_tooltip(key):
    """Muestra un tooltip discreto"""
    if key in TOOLTIPS:
        st.markdown(f'<span style="color: var(--text-secondary); font-size: 0.85rem; font-style: italic;">{TOOLTIPS[key]}</span>', unsafe_allow_html=True)

# ==========================================
# ESTADO DE SESIÓN
# ==========================================
if 'data_seo' not in st.session_state:
    st.session_state.data_seo = []
if 'global_words' not in st.session_state:
    st.session_state.global_words = []
if 'gap_analysis' not in st.session_state:
    st.session_state.gap_analysis = {}
if 'show_tooltips' not in st.session_state:
    st.session_state.show_tooltips = False

# ==========================================
# FUNCIONES DE SCRAPING (sin cambios)
# ==========================================
def get_serp_urls(keyword, num=10):
    """Obtiene URLs del SERP de Google"""
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
    }
    query = keyword.replace(' ', '+')
    url = f"https://www.google.com/search?q={query}&num={num+5}"
    
    try:
        resp = requests.get(url, headers=headers, timeout=10, verify=False)
        soup = BeautifulSoup(resp.text, 'html.parser')
        urls = []
        
        for link in soup.select('a'):
            href = link.get('href', '')
            if '/url?q=' in href:
                clean = href.split('/url?q=')[1].split('&')[0]
                if clean.startswith('http') and 'google.com' not in clean:
                    urls.append(clean)
        
        return urls[:num]
    except:
        return []

def infer_intent(title, meta, content):
    """Inferir intención de búsqueda"""
    text = f"{title} {meta} {content}".lower()
    
    info_signals = ['qué es', 'cómo', 'guía', 'tutorial', 'aprende', 'what is', 'how to', 'guide', 'learn']
    commercial_signals = ['mejor', 'top', 'comparar', 'vs', 'review', 'opiniones', 'best', 'compare', 'versus']
    transactional_signals = ['comprar', 'precio', 'oferta', 'descuento', 'tienda', 'buy', 'price', 'deal', 'discount', 'shop']
    
    scores = {
        'Informational': sum(1 for s in info_signals if s in text),
        'Commercial': sum(1 for s in commercial_signals if s in text),
        'Transactional': sum(1 for s in transactional_signals if s in text)
    }
    
    return max(scores, key=scores.get) if max(scores.values()) > 0 else 'Informational'

def extract_seo_data(url, keyword):
    """Extrae datos SEO de una URL"""
    headers = {
        'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36'
    }
    
    try:
        resp = requests.get(url, headers=headers, timeout=15, verify=False)
        soup = BeautifulSoup(resp.text, 'html.parser')
        
        # Extraer contenido principal
        full_text = trafilatura.extract(resp.text) or ""
        
        # Title y meta
        title = soup.find('title')
        title = title.get_text().strip() if title else ""
        
        meta_desc = soup.find('meta', attrs={'name': 'description'})
        meta_desc = meta_desc.get('content', '').strip() if meta_desc else ""
        
        # H1 y H2
        h1_list = [h.get_text().strip() for h in soup.find_all('h1')]
        h2_list = [h.get_text().strip() for h in soup.find_all('h2')]
        
        # Word count
        words = re.findall(r'\b\w+\b', full_text.lower())
        word_count = len(words)
        
        # Keyword density
        kw_lower = keyword.lower()
        kw_density = sum(1 for w in words if kw_lower in w)
        
        # Enlaces
        all_links = soup.find_all('a', href=True)
        internal_links = [l for l in all_links if urlparse(url).netloc in l['href']]
        external_links = [l for l in all_links if urlparse(url).netloc not in l['href'] and l['href'].startswith('http')]
        external_dofollow = [l for l in external_links if not l.get('rel') or 'nofollow' not in l.get('rel')]
        
        # Imágenes
        images = soup.find_all('img')
        images_with_alt = [img for img in images if img.get('alt')]
        videos = soup.find_all(['video', 'iframe'])
        
        # Schemas
        schemas = []
        for script in soup.find_all('script', type='application/ld+json'):
            try:
                data = json.loads(script.string)
                if '@type' in data:
                    schemas.append(data['@type'])
            except:
                pass
        
        # SEO On-Page
        seo_onpage = {
            'kw_in_title': kw_lower in title.lower(),
            'kw_in_meta': kw_lower in meta_desc.lower(),
            'kw_in_h1': any(kw_lower in h.lower() for h in h1_list),
            'kw_in_strong': bool(soup.find('strong', string=re.compile(kw_lower, re.I))),
            'title_length': len(title),
            'meta_length': len(meta_desc)
        }
        
        # Readability
        try:
            readability = flesch_reading_ease(full_text)
        except:
            readability = 0
        
        # DA Proxy (simulado)
        da_proxy = min(100, len(external_dofollow) * 5 + len(internal_links) * 2)
        
        # Intención
        intencion = infer_intent(title, meta_desc, full_text)
        
        return {
            'title': title,
            'meta_desc': meta_desc,
            'h1': h1_list,
            'h2': h2_list,
            'word_count': word_count,
            'kw_density': kw_density,
            'full_text': full_text,
            'enlaces': {
                'internal': len(internal_links),
                'external': len(external_links),
                'external_dofollow': len(external_dofollow)
            },
            'media': {
                'images': len(images),
                'images_with_alt': len(images_with_alt),
                'videos': len(videos),
                'total': len(images) + len(videos)
            },
            'schemas': schemas,
            'seo_onpage': seo_onpage,
            'readability': round(readability, 1),
            'da_proxy': da_proxy,
            'intencion': intencion
        }
    except Exception as e:
        return None

def get_ngrams(words, n):
    """Extrae n-gramas"""
    if len(words) < n:
        return []
    return [' '.join(words[i:i+n]) for i in range(len(words)-n+1)]

def calcular_gap_analysis(data, keyword):
    """Calcula el gap analysis"""
    if not data:
        return {}
    
    # Benchmarks
    word_avg = np.mean([d['Palabras'] for d in data])
    h2_avg = np.mean([len(d.get('H2_list', d.get('h2', []))) for d in data])
    internal_avg = np.mean([d.get('Enlaces', d.get('enlaces', {})).get('internal', 0) for d in data])
    external_avg = np.mean([d.get('Enlaces', d.get('enlaces', {})).get('external', 0) for d in data])
    media_avg = np.mean([d.get('Media', d.get('media', {})).get('total', 0) for d in data])
    
    # H2s críticos
    all_h2s = []
    for item in data:
        h2_list = item.get('H2_list', item.get('h2', []))
        all_h2s.extend([h.lower().strip() for h in h2_list])
    
    h2_counter = Counter(all_h2s)
    threshold = len(data) * 0.5
    h2s_criticos = {h2: count for h2, count in h2_counter.items() if count >= threshold}
    
    return {
        'coverage_benchmark': {
            'word_avg': word_avg,
            'h2_avg': h2_avg,
            'internal_links_avg': internal_avg,
            'external_links_avg': external_avg,
            'media_avg': media_avg
        },
        'h2s_criticos': h2s_criticos
    }

# ==========================================
# EXPORTACIÓN PROFESIONAL A EXCEL
# ==========================================
def export_to_excel(df, gap_analysis, keyword):
    """Exporta los resultados a Excel con formato profesional"""
    output = io.BytesIO()
    wb = Workbook()
    
    # Estilos
    header_fill = PatternFill(start_color="FF6B00", end_color="FF6B00", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    border = Border(
        left=Side(style='thin', color='2d3250'),
        right=Side(style='thin', color='2d3250'),
        top=Side(style='thin', color='2d3250'),
        bottom=Side(style='thin', color='2d3250')
    )
    
    # Sheet 1: Overview
    ws1 = wb.active
    ws1.title = "Overview"
    
    # Headers
    ws1['A1'] = f"SERP Analysis Report - {keyword}"
    ws1['A1'].font = Font(bold=True, size=14)
    ws1['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    
    # Métricas principales
    ws1['A4'] = "Key Metrics"
    ws1['A4'].font = Font(bold=True, size=12)
    
    if gap_analysis and 'coverage_benchmark' in gap_analysis:
        bench = gap_analysis['coverage_benchmark']
        metrics = [
            ['Metric', 'Average'],
            ['Word Count', f"{bench['word_avg']:.0f}"],
            ['H2 Headings', f"{bench['h2_avg']:.1f}"],
            ['Internal Links', f"{bench['internal_links_avg']:.1f}"],
            ['External Links', f"{bench['external_links_avg']:.1f}"],
            ['Media Elements', f"{bench['media_avg']:.1f}"],
        ]
        
        for idx, row in enumerate(metrics, start=5):
            ws1[f'A{idx}'] = row[0]
            ws1[f'B{idx}'] = row[1]
            if idx == 5:
                ws1[f'A{idx}'].fill = header_fill
                ws1[f'B{idx}'].fill = header_fill
                ws1[f'A{idx}'].font = header_font
                ws1[f'B{idx}'].font = header_font
    
    # Sheet 2: Detailed Data
    ws2 = wb.create_sheet("Competitor Details")
    
    # Columnas simplificadas para Excel
    export_df = df[['Pos', 'URL', 'Título', 'Palabras', 'Intención', 'Menciones KW', 'DA_Proxy']].copy()
    
    for r_idx, row in enumerate(dataframe_to_rows(export_df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            cell = ws2.cell(row=r_idx, column=c_idx, value=value)
            cell.border = border
            if r_idx == 1:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center')
    
    # Ajustar anchos
    for ws in [ws1, ws2]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(output)
    output.seek(0)
    return output

# ==========================================
# INTERFAZ PRINCIPAL
# ==========================================

# HEADER
st.markdown("""
<div class="main-header">
    <h1>SERP X-Ray 360</h1>
    <div class="subtitle">Professional SEO Analysis Platform</div>
</div>
""", unsafe_allow_html=True)

# SIDEBAR
with st.sidebar:
    st.markdown("### Configuration")
    
    keyword = st.text_input(
        "Target Keyword",
        value="lamparas cataliticas",
        help="Enter the keyword you want to analyze"
    )
    
    num_competitors = st.number_input(
        "Competitors to Analyze",
        min_value=3,
        max_value=20,
        value=10,
        help="Number of top-ranking pages to analyze"
    )
    
    st.markdown("---")
    
    st.markdown("### Options")
    st.session_state.show_tooltips = st.checkbox(
        "Show Tooltips",
        value=st.session_state.show_tooltips,
        help="Display helpful tooltips throughout the interface"
    )
    
    st.markdown("---")
    
    if st.button("🚀 Run Analysis", use_container_width=True):
        if keyword:
            with st.spinner("Analyzing SERP..."):
                urls = get_serp_urls(keyword, num_competitors)
                
                if urls:
                    final_data = []
                    global_words = []
                    count_valid = 0
                    
                    progress_bar = st.progress(0)
                    status = st.status("Scraping competitors...", expanded=True)
                    
                    for idx, url in enumerate(urls):
                        status.write(f"Analyzing: {url[:60]}...")
                        data = extract_seo_data(url, keyword)
                        
                        if data:
                            words = re.findall(r'\b\w+\b', data.get('full_text', '').lower())
                            global_words.extend(words)
                            
                            final_data.append({
                                "Pos": count_valid + 1,
                                "URL": url,
                                "Título": data['title'],
                                "Meta Desc": data['meta_desc'],
                                "Palabras": data['word_count'],
                                "Intención": data['intencion'],
                                "Menciones KW": data['kw_density'],
                                "H1_list": data.get('h1', []),
                                "H2_list": data.get('h2', []),
                                "Contenido": data.get('full_text', '')[:1000],
                                "Enlaces": data.get('enlaces', {}),
                                "SEO_Onpage": data.get('seo_onpage', {}),
                                "Schemas": data.get('schemas', []),
                                "Media": data.get('media', {}),
                                "Readability": data.get('readability', 'N/A'),
                                "DA_Proxy": data.get('da_proxy', 0),
                                **data
                            })
                            count_valid += 1
                            progress_bar.progress(count_valid / num_competitors)
                    
                    if final_data:
                        st.session_state.data_seo = final_data
                        st.session_state.global_words = global_words
                        st.session_state.gap_analysis = calcular_gap_analysis(final_data, keyword)
                        status.update(label="Analysis Complete", state="complete")
                        st.success("Analysis completed successfully!")

# ==========================================
# RESULTADOS
# ==========================================
if st.session_state.data_seo:
    data = st.session_state.data_seo
    df = pd.DataFrame(data)
    
    # Botón de exportación
    st.markdown("### Export Results")
    col1, col2, col3 = st.columns([1, 1, 2])
    
    with col1:
        excel_file = export_to_excel(df, st.session_state.gap_analysis, keyword)
        st.download_button(
            label="Download Excel Report",
            data=excel_file,
            file_name=f"serp_analysis_{keyword.replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True
        )
    
    st.markdown("---")
    
    tab1, tab2, tab3 = st.tabs([
        "Overview", 
        "Gap Analysis", 
        "Competitor Details"
    ])
    
    with tab1:
        st.markdown("### Key Metrics")
        
        c1, c2, c3, c4, c5 = st.columns(5)
        
        word_avg = int(df['Palabras'].mean())
        c1.metric("Avg Word Count", word_avg)
        if st.session_state.show_tooltips:
            with c1:
                show_tooltip("word_count")
        
        c2.metric("Search Intent", df['Intención'].mode()[0] if not df['Intención'].empty else "N/A")
        
        internal_avg = int(df.apply(lambda x: x.get('Enlaces', x.get('enlaces', {})).get('internal', 0) if isinstance(x.get('Enlaces', x.get('enlaces', {})), dict) else 0, axis=1).mean())
        c3.metric("Avg Internal Links", internal_avg)
        if st.session_state.show_tooltips:
            with c3:
                show_tooltip("internal_links")
        
        media_avg = df.apply(lambda x: x.get('Media', x.get('media', {})).get('total', 0) if isinstance(x.get('Media', x.get('media', {})), dict) else 0, axis=1).mean()
        c4.metric("Avg Media Elements", f"{media_avg:.1f}")
        if st.session_state.show_tooltips:
            with c4:
                show_tooltip("images")
        
        c5.metric("Avg Domain Authority", int(df['DA_Proxy'].mean()))
        if st.session_state.show_tooltips:
            with c5:
                show_tooltip("da_proxy")
        
        st.markdown("---")
        
        # Secondary Keywords
        st.markdown("### Most Common Secondary Keywords")
        if st.session_state.show_tooltips:
            st.markdown('<div class="info-box">These are the 2-word phrases that appear most frequently across top-ranking content. Include these naturally in your content.</div>', unsafe_allow_html=True)
        
        bigrams = get_ngrams(st.session_state.global_words, 2)
        if bigrams:
            chart_data = pd.DataFrame(Counter(bigrams).most_common(12), 
                                     columns=['Phrase', 'Frequency']).set_index('Phrase')
            st.bar_chart(chart_data, color="#FF6B00")
    
    with tab2:
        st.markdown("### Content Strategy Recommendations")
        
        if st.session_state.gap_analysis:
            gap = st.session_state.gap_analysis
            bench = gap['coverage_benchmark']
            
            # Action Checklist
            st.markdown("#### Recommended Content Structure")
            
            checklist_items = [
                f"Write at least **{int(bench['word_avg'] * 1.2):,} words** (20% above average of {int(bench['word_avg']):,})",
                f"Include minimum **{int(bench['h2_avg'])} H2 sections** for proper content structure",
                f"Add **{int(bench['internal_links_avg'])} internal links** to relevant pages",
                f"Reference **{int(bench['external_links_avg'])} authoritative external sources**",
                f"Include **{int(bench['media_avg'])} media elements** (images/videos with optimized ALT text)",
            ]
            
            for item in checklist_items:
                st.markdown(f"- {item}")
            
            st.markdown("---")
            
            # Critical H2s
            st.markdown("#### Critical Content Sections")
            if gap['h2s_criticos']:
                if st.session_state.show_tooltips:
                    st.markdown('<div class="info-box">These H2 headings appear in more than 50% of competing pages. Search engines expect to see these topics covered.</div>', unsafe_allow_html=True)
                
                h2_df = pd.DataFrame(
                    [(h, count) for h, count in gap['h2s_criticos'].items()],
                    columns=['H2 Section', 'Appears in X Competitors']
                ).sort_values('Appears in X Competitors', ascending=False)
                st.dataframe(h2_df, use_container_width=True)
                
                st.markdown('<div class="success-box"><strong>Action:</strong> Create original content for each of these sections.</div>', unsafe_allow_html=True)
            else:
                st.info("No consistently repeated H2s found (competitors use varied structures)")
    
    with tab3:
        st.markdown("### Detailed Competitor Analysis")
        
        for item in data:
            titulo_corto = (item['Título'][:80] + '..') if len(item['Título']) > 80 else item['Título']
            
            with st.expander(f"#{item['Pos']} | {titulo_corto}", expanded=False):
                
                # Main metrics
                col1, col2, col3 = st.columns(3)
                
                with col1:
                    st.markdown("#### Basic Metrics")
                    st.markdown(f"**Word Count:** {item['Palabras']:,}")
                    st.markdown(f"**Intent:** {item['Intención']}")
                    st.markdown(f"**DA Score:** {item['DA_Proxy']}")
                    st.markdown(f"**Readability:** {item['Readability']}")
                
                with col2:
                    st.markdown("#### Links")
                    enlaces = item.get('enlaces', {})
                    st.markdown(f"**Internal:** {enlaces.get('internal', 0)}")
                    st.markdown(f"**External:** {enlaces.get('external', 0)}")
                    st.markdown(f"**External Dofollow:** {enlaces.get('external_dofollow', 0)}")
                
                with col3:
                    st.markdown("#### Media")
                    media = item.get('media', {})
                    st.markdown(f"**Images:** {media.get('images', 0)}")
                    st.markdown(f"**Videos:** {media.get('videos', 0)}")
                    st.markdown(f"**Images with ALT:** {media.get('images_with_alt', 0)}")
                
                # SEO On-Page
                seo = item.get('SEO_Onpage', item.get('seo_onpage', {}))
                if seo:
                    st.markdown("#### On-Page SEO")
                    
                    checks = [
                        ("Keyword in Title", seo.get('kw_in_title', False)),
                        ("Keyword in Meta Description", seo.get('kw_in_meta', False)),
                        ("Keyword in H1", seo.get('kw_in_h1', False)),
                        ("Keyword in Bold Text", seo.get('kw_in_strong', False))
                    ]
                    
                    col_a, col_b = st.columns(2)
                    for idx, (check_name, has_it) in enumerate(checks):
                        badge_class = "badge-success" if has_it else "badge-danger"
                        badge_text = "YES" if has_it else "NO"
                        target_col = col_a if idx % 2 == 0 else col_b
                        with target_col:
                            st.markdown(f'<span class="badge {badge_class}">{badge_text}</span> {check_name}', unsafe_allow_html=True)
                    
                    st.markdown(f"**Title Length:** {seo.get('title_length', 0)} chars")
                    st.markdown(f"**Meta Length:** {seo.get('meta_length', 0)} chars")
                
                # H2s
                h2_list = item.get('H2_list', item.get('h2', []))
                if h2_list:
                    st.markdown("#### H2 Structure")
                    for h2 in h2_list[:10]:
                        st.markdown(f"- {h2}")
                
                # Schemas
                schemas = item.get('Schemas', item.get('schemas', []))
                if schemas:
                    st.markdown(f"#### Schema Types: {', '.join(schemas)}")

# ==========================================
# FOOTER
# ==========================================
st.markdown("---")
st.markdown("""
<div style='text-align: center; padding: 2rem 0; color: var(--text-secondary);'>
    <p style='font-size: 0.9rem;'><strong>SERP X-Ray 360</strong> | Professional SEO Analysis Platform</p>
    <p style='font-size: 0.75rem;'>Enterprise-grade SERP intelligence © 2025</p>
</div>
""", unsafe_allow_html=True)